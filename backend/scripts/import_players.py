"""Erase and reload the `players` table from a fantacalcio.it Quotazioni xlsx.

Reads the `Tutti` sheet only (the `Ceduti` sheet — transferred-out players —
is intentionally skipped). Unknown Team or MantraRole values raise loudly so
that a new value in the source file becomes a visible schema-change signal.

Usage (CLI):
    uv run python -m backend.scripts.import_players \\
        data/Quotazioni_Fantacalcio_Stagione_2025_26.xlsx

Library:
    parse_players(source) -> list[Player]   # source: Path | str | bytes | file-like
    write_players(rows)   -> int            # TRUNCATE + INSERT in one transaction
    import_xlsx(source)   -> int            # parse then write
"""
import sys
from io import BytesIO
from pathlib import Path
from typing import IO, Union

from openpyxl import load_workbook
from sqlalchemy import text

from backend.app.db import SessionLocal
from backend.app.models import MantraRole, Player

SHEET = "Tutti"
EXPECTED_TEAM_COUNT = 20  # Serie A has exactly 20 clubs

Source = Union[Path, str, bytes, IO[bytes]]


def _parse_mantra_roles(rm: object) -> list[MantraRole]:
    return [MantraRole(token.strip()) for token in str(rm).split(";") if token.strip()]


def _as_workbook_source(source: Source):
    if isinstance(source, (str, Path)):
        return source
    if isinstance(source, (bytes, bytearray)):
        return BytesIO(bytes(source))
    return source


def parse_players(source: Source) -> list[Player]:
    """Parse an xlsx (path, bytes, or file-like) into Player rows.

    Raises:
        openpyxl.utils.exceptions.InvalidFileException: not a valid xlsx.
        KeyError: required sheet missing.
        ValueError: unknown Team or MantraRole value.
    """
    wb = load_workbook(_as_workbook_source(source), data_only=True)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet {SHEET!r} not found in workbook")

    ws = wb[SHEET]
    rows: list[Player] = []
    for row_idx, raw in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        pid = raw[0]
        # Skip blank rows and rows where the first cell isn't a numeric Id.
        # This tolerates extra title/footer/totals rows that some xlsx exports add.
        if pid is None:
            continue
        try:
            pid_int = int(pid)
        except (TypeError, ValueError):
            continue
        try:
            rows.append(
                Player(
                    id=pid_int,
                    name=str(raw[3]).strip(),
                    team=str(raw[4]).strip(),
                    mantra_roles=_parse_mantra_roles(raw[2]),
                    # raw[8] = "Qt.A M" (Mantra evaluation)
                    fanta_evaluation=raw[8],
                    # raw[12] = "FVM M", stored as-is.
                    fanta_market_value=raw[12],
                )
            )
        except ValueError as exc:
            raise ValueError(f"row {row_idx} ({raw!r}): {exc}") from exc

    teams = {p.team for p in rows}
    if len(teams) != EXPECTED_TEAM_COUNT:
        raise ValueError(
            f"Expected exactly {EXPECTED_TEAM_COUNT} distinct teams in the file, "
            f"got {len(teams)}: {sorted(teams)}"
        )
    return rows


def write_players(rows: list[Player]) -> int:
    """Replace the players table with `rows` atomically (TRUNCATE + INSERT)."""
    with SessionLocal() as session:
        with session.begin():
            session.execute(text("TRUNCATE TABLE players RESTART IDENTITY"))
            session.add_all(rows)
    return len(rows)


def import_xlsx(source: Source) -> int:
    """Parse + write. Returns the inserted row count."""
    rows = parse_players(source)
    return write_players(rows)


def main() -> None:
    if len(sys.argv) != 2:
        print("Usage: python -m backend.scripts.import_players <xlsx>", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)
    n = import_xlsx(path)
    print(f"Imported {n} players.")


if __name__ == "__main__":
    main()
